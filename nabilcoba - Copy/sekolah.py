import cv2
import sqlite3
import pathlib
import os
import tkinter as tk
from datetime import datetime, date, timedelta
from PIL import Image, ImageTk
import numpy as np
import face_recognition
import pyttsx3  # Import pyttsx3 for text-to-speech
import openpyxl  # Import openpyxl for Excel export

# Paths
current_dir = pathlib.Path(__file__).parent
img_dir = current_dir.joinpath('img')
db_path = current_dir.joinpath('faces.db')

# Create directories if not exist
if not img_dir.exists():
    os.makedirs(img_dir)

# Connect to SQLite database
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Create table for storing face data
cursor.execute('''
CREATE TABLE IF NOT EXISTS faces (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    role TEXT,
    image_path TEXT,
    encoding BLOB,
    log TEXT
)
''')
conn.commit()

detected_faces = {}

def recognize_face(face_encoding):
    cursor.execute("SELECT * FROM faces")
    rows = cursor.fetchall()
    for row in rows:
        known_encoding = np.frombuffer(row[4], dtype=np.float64)
        matches = face_recognition.compare_faces([known_encoding], face_encoding, tolerance=0.6)
        if True in matches:
            return row
    return None

def save_identity(name, face, role, face_encoding):
    timestamp = datetime.now().timestamp()
    img_path = img_dir.joinpath(f'{int(timestamp)}.jpg')
    cv2.imwrite(str(img_path), face)  # Convert img_path to string

    encoding_blob = face_encoding.tobytes()
    cursor.execute('''
    INSERT INTO faces (name, role, image_path, encoding, log) VALUES (?, ?, ?, ?, ?)
    ''', (name, role, str(img_path), encoding_blob, '[]'))
    conn.commit()

def register_new_face(frame, face, face_encoding):
    root = tk.Tk()
    root.title("Daftarkan Wajah")

    tk.Label(root, text='Masukkan nama:').pack()
    name_entry = tk.Entry(root)
    name_entry.pack()

    tk.Label(root, text='Masukkan role:').pack()
    role_entry = tk.Entry(root)
    role_entry.pack()

    # Show the face to be saved
    face_img = Image.fromarray(cv2.cvtColor(face, cv2.COLOR_BGR2RGB))
    face_img = ImageTk.PhotoImage(image=face_img)
    tk.Label(root, image=face_img).pack()

    def save():
        name = name_entry.get()
        role = role_entry.get()
        save_identity(name, face, role, face_encoding)
        engine.say(f"Welcome {name}")
        engine.runAndWait()
        root.destroy()

    tk.Button(root, text='Simpan', command=save).pack()
    root.mainloop()

def log_attendance(entry_id, action):
    today = date.today().isoformat()
    now = datetime.now().isoformat()
    
    cursor.execute("SELECT log FROM faces WHERE id = ?", (entry_id,))
    log = eval(cursor.fetchone()[0])  # Convert string representation of list back to list

    if action == "datang":
        if len(log) == 0 or log[-1]['date'] != today:
            log.append({
                'date': today,
                'attendance_time': now,
                'go_home_time': None,
                'duration': None
            })
            print(f"Datang: {entry_id} at {now}")
    elif action == "pulang":
        if log[-1]['date'] == today and log[-1]['go_home_time'] is None:
            log[-1]['go_home_time'] = now
            # Calculate duration
            arrival_time = datetime.fromisoformat(log[-1]['attendance_time'])
            departure_time = datetime.fromisoformat(now)
            duration = departure_time - arrival_time
            log[-1]['duration'] = str(duration)
            print(f"Pulang: {entry_id} at {now}")
            print(f"Durasi: {duration}")

    cursor.execute("UPDATE faces SET log = ? WHERE id = ?", (str(log), entry_id))
    conn.commit()

# Initialize text-to-speech engine
engine = pyttsx3.init()

def export_to_excel():
    cursor.execute("SELECT * FROM faces")
    rows = cursor.fetchall()

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Face Data"

    # Write the header row
    headers = ["ID", "Name", "Role", "Image Path", "Encoding", "Log"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Write data rows
    for row_num, row in enumerate(rows, 2):
        sheet.cell(row=row_num, column=1, value=row[0])
        sheet.cell(row=row_num, column=2, value=row[1])
        sheet.cell(row=row_num, column=3, value=row[2])
        sheet.cell(row=row_num, column=4, value=row[3])
        sheet.cell(row=row_num, column=5, value=str(np.frombuffer(row[4], dtype=np.float64)))
        sheet.cell(row=row_num, column=6, value=row[5])

    # Save the workbook to a file
    excel_path = current_dir.joinpath('faces_data.xlsx')
    workbook.save(excel_path)
    print(f"Data telah diekspor ke {excel_path}")

cap = cv2.VideoCapture(0)

running = True
while running:
    ret, frame = cap.read()
    if not ret:
        break

    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    face_locations = face_recognition.face_locations(rgb_frame)
    face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        face_img = frame[top:bottom, left:right]
        
        # Generate a unique key for the detected face based on its position
        face_key = (left, top, right, bottom)
        
        entry = recognize_face(face_encoding)
        if entry:
            current_time = datetime.now()
            entry_id = entry[0]
            
            # Check if the face was detected recently
            last_detected = detected_faces.get(entry_id, None)
            
            if last_detected is None or (current_time - last_detected) >= timedelta(minutes=1):
                if last_detected is None:
                    print(f"Datang: {entry[1]} ({entry[2]}) at {current_time}")
                    log_attendance(entry_id, "datang")
                    engine.say(f"Welcome {entry[1]}")
                    engine.runAndWait()
                else:
                    print(f"Pulang: {entry[1]} ({entry[2]}) at {current_time}")
                    log_attendance(entry_id, "pulang")
                    engine.say(f"Goodbye {entry[1]}, be careful on the road")
                    engine.runAndWait()
                
                detected_faces[entry_id] = current_time
                name_and_role = f"{entry[1]} ({entry[2]})"
                cv2.putText(frame, name_and_role, (left, top-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)

        else:
            register_new_face(frame, face_img, face_encoding)

        cv2.rectangle(frame, (left, top), (right, bottom), (255, 0, 0), 2)

    # Display real-time clock on the window
    cv2.putText(frame, f'Time: {datetime.now().strftime("%H:%M:%S")}', (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
    cv2.putText(frame, 'Tekan [q] untuk keluar', (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
    cv2.imshow('Camera', frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
conn.close()

# Call the export_to_excel function when you want to export data
export_to_excel()
