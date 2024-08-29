import cv2
import sqlite3
import pathlib
import os
import tkinter as tk
from tkinter import messagebox
from datetime import datetime, date
from PIL import Image, ImageTk
import numpy as np
import face_recognition
from openpyxl import Workbook, load_workbook
import pyttsx3  # Untuk text-to-speech

# Paths
current_dir = pathlib.Path(__file__).parent
img_dir = current_dir.joinpath('img')
screenshots_dir = current_dir.joinpath('screenshots')
morning_screenshots_dir = screenshots_dir.joinpath('morning')
afternoon_screenshots_dir = screenshots_dir.joinpath('afternoon')
db_path = current_dir.joinpath('faces.db')
spreadsheet_path = current_dir.joinpath('face_data.xlsx')

# Create directories if not exist
for dir_path in [img_dir, screenshots_dir, morning_screenshots_dir, afternoon_screenshots_dir]:
    if not dir_path.exists():
        os.makedirs(dir_path)

# Create or load the spreadsheet
if spreadsheet_path.exists():
    workbook = load_workbook(spreadsheet_path)
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Face Data"
    sheet.append(["Name", "Role", "Image Path", "Log", "Arrival Time", "Return Time", "Duration"])  # Header row
    workbook.save(spreadsheet_path)

# Connect to SQLite database
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Create table for storing face data
cursor.execute('''
CREATE TABLE IF NOT EXISTS faces (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    role TEXT,
    image_path TEXT,
    encoding BLOB,
    log TEXT
)
''')
conn.commit()

# Initialize text-to-speech engine
engine = pyttsx3.init()

def save_to_spreadsheet(name, role, img_path, log, arrival_time, return_time, duration):
    workbook = load_workbook(spreadsheet_path)
    sheet = workbook.active

    # Append the data to the spreadsheet
    sheet.append([name, role, img_path, log, arrival_time, return_time, duration])
    
    # Save the workbook
    workbook.save(spreadsheet_path)
    print("Data saved to Excel.")

def recognize_face(face_encoding):
    cursor.execute("SELECT * FROM faces")
    rows = cursor.fetchall()
    for row in rows:
        known_encoding = np.frombuffer(row[4], dtype=np.float64)
        matches = face_recognition.compare_faces([known_encoding], face_encoding, tolerance=0.4)
        if True in matches:
            return row
    return None

def save_identity(name, face, role, face_encoding):
    timestamp = datetime.now().timestamp()
    img_path = img_dir.joinpath(f'{int(timestamp)}.jpg')
    cv2.imwrite(str(img_path), face)

    encoding_blob = face_encoding.tobytes()
    cursor.execute('''
    INSERT INTO faces (name, role, image_path, encoding, log) VALUES (?, ?, ?, ?, ?)
    ''', (name, role, str(img_path), encoding_blob, '[]'))
    conn.commit()

    save_to_spreadsheet(name, role, str(img_path), '[]', '', '', '')

def register_new_face(frame, face, face_encoding):
    global register_mode
    register_mode = True

    def save():
        name = name_entry.get()
        role = role_entry.get()
        if name and role:
            save_identity(name, face, role, face_encoding)
            root.destroy()
            global register_mode
            register_mode = False  # Reset the registration mode after saving
        else:
            messagebox.showwarning("Input Error", "Name and Role are required")

    root = tk.Tk()
    root.title("Daftarkan Wajah")

    tk.Label(root, text='Masukkan nama:').pack()
    name_entry = tk.Entry(root)
    name_entry.pack()

    tk.Label(root, text='Masukkan role:').pack()
    role_entry = tk.Entry(root)
    role_entry.pack()

    # Convert face image to display
    face_img = Image.fromarray(cv2.cvtColor(face, cv2.COLOR_BGR2RGB))
    face_img = ImageTk.PhotoImage(image=face_img)
    tk.Label(root, image=face_img).pack()

    tk.Button(root, text='Simpan', command=save).pack()
    root.mainloop()

def log_attendance(entry_id):
    today = date.today().isoformat()
    now = datetime.now().isoformat()
    
    cursor.execute("SELECT log FROM faces WHERE id = ?", (entry_id,))
    log = eval(cursor.fetchone()[0])

    if len(log) == 0 or log[-1]['date'] != today:
        log.append({
            'date': today,
            'arrival_time': now,
            'return_time': None
        })
        # Announce arrival
        entry_name = cursor.execute("SELECT name FROM faces WHERE id = ?", (entry_id,)).fetchone()[0]
        engine.say(f"Welcome {entry_name} in Praxis High School")
        engine.runAndWait()
    else:
        log[-1]['return_time'] = now
        # Calculate duration
        arrival_time = datetime.fromisoformat(log[-1]['arrival_time'])
        return_time = datetime.fromisoformat(log[-1]['return_time'])
        duration = return_time - arrival_time
        duration_str = str(duration)

        # Announce departure
        entry_name = cursor.execute("SELECT name FROM faces WHERE id = ?", (entry_id,)).fetchone()[0]
        engine.say(f"Goodbye {entry_name}, be careful on road.")
        engine.runAndWait()

        # Update the spreadsheet
        save_to_spreadsheet(entry_name, "", "", str(log), log[-1]['arrival_time'], log[-1]['return_time'], duration_str)

    cursor.execute("UPDATE faces SET log = ? WHERE id = ?", (str(log), entry_id))
    conn.commit()

def get_screenshot_folder():
    now = datetime.now()
    current_time = now.time()
    morning_start = datetime.strptime('07:30:00', '%H:%M:%S').time()
    morning_end = datetime.strptime('12:00:00', '%H:%M:%S').time()
    afternoon_start = datetime.strptime('14:00:00', '%H:%M:%S').time()
    night_end = datetime.strptime('00:00:00', '%H:%M:%S').time()

    if morning_start <= current_time < morning_end:
        return morning_screenshots_dir
    elif afternoon_start <= current_time or current_time < night_end:
        return afternoon_screenshots_dir
    else:
        return None

cap = cv2.VideoCapture(0)

# Set lower resolution to speed up processing
cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

screenshot_taken_for_face = {}  # Dictionary to track if screenshot is taken for a specific face
register_mode = False
frame_count = 0
frame_skip = 5  # Only detect faces every 5 frames

while True:
    ret, frame = cap.read()
    if not ret:
        break

    frame = cv2.convertScaleAbs(frame, alpha=1.3, beta=30)

    frame_count += 1
    if frame_count % frame_skip != 0:
        cv2.imshow('Frame', frame)
        key = cv2.waitKey(1) & 0xFF
        if key == ord('q'):
            break
        continue

    # Downscale the frame for faster processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

    # Use the 'hog' model for faster face detection
    face_locations = face_recognition.face_locations(rgb_small_frame, model="hog")
    
    # Process only the largest face for performance reasons
    if face_locations:
        largest_face_location = max(face_locations, key=lambda rect: (rect[2] - rect[0]) * (rect[1] - rect[3]))
        # Scale back the face location to the original frame size
        top, right, bottom, left = [v * 4 for v in largest_face_location]
        face_encodings = face_recognition.face_encodings(frame, [(top, right, bottom, left)], num_jitters=0)

        for face_encoding in face_encodings:
            face_img = frame[top:bottom, left:right]
            face_key = tuple(np.round(face_encoding, decimals=5))  # Round to avoid floating-point issues

            if register_mode:
                # Skip recognition during registration mode
                cv2.putText(frame, "Mode Registrasi Aktif", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 255), 2)
                continue

            entry = recognize_face(face_encoding)
            if entry:
                log_attendance(entry[0])
                name_and_role = f"{entry[1]} ({entry[2]})"
                cv2.putText(frame, name_and_role, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

                if face_key not in screenshot_taken_for_face:
                    screenshot_folder = get_screenshot_folder()
                    if screenshot_folder:
                        screenshot_path = screenshot_folder.joinpath(f'{int(datetime.now().timestamp())}.jpg')
                        cv2.imwrite(str(screenshot_path), frame)
                        screenshot_taken_for_face[face_key] = screenshot_path
            else:
                register_new_face(frame, face_img, face_encoding)

    cv2.imshow('Frame', frame)
    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
conn.close()
